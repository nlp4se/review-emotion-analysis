import os
import sys
import pandas as pd
import numpy as np
from sklearn.metrics import cohen_kappa_score
from openpyxl import Workbook
from openpyxl.styles import Font
import logging
import seaborn as sns
import matplotlib.pyplot as plt
import argparse

# Fixed list of all possible annotators in specified order
#all_annotators = ['gpt-4o', 'mistral-large-2411', 'gemini-2-0-flash']
#all_annotators = ['gpt-4o-1','gpt-4o-2','gpt-4o-3']
#all_annotators = ['mistral-large-2411-1','mistral-large-2411-2','mistral-large-2411-3']
#all_annotators = ['gemini-2-0-flash-1','gemini-2-0-flash-2','gemini-2-0-flash-3']
all_annotators = ['Ann_1', 'Ann_2', 'Ann_3', 'Ann_4', 'Ann_5', 
                  'agreement_Ann_1-Ann_2-Ann_3-Ann_4-Ann_5', 
                  'agreement_gpt-4o-1-gpt-4o-2-gpt-4o-3',
                  'agreement_mistral-large-2411-1-mistral-large-2411-2-mistral-large-2411-3',
                  'agreement_gemini-2-0-flash-1-gemini-2-0-flash-2-gemini-2-0-flash-3', 
                  'agreement_gemini-2-0-flash-1-gemini-2-0-flash-2-gemini-2-0-flash-3-gpt-4o-1-gpt-4o-2-gpt-4o-3-mistral-large-2411-1-mistral-large-2411-2-mistral-large-2411-3']

heatmap_tags = ['Ann_1', 'Ann_2', 'Ann_3', 'Ann_4', 'Ann_5', 'Ann_A', 'GPT_A', 'Mistral_A', 'Gemini_A', 'LLM_A']

def get_annotation_data(folder_path, exclude_iterations=None):
    """Extract annotation data from all iteration folders."""
    logging.info(f"Starting to process folder: {folder_path}")
    iterations_data = {}
    
    try:
        # Find all iteration folders
        for item in os.listdir(folder_path):
            if item.startswith('iteration_'):
                iteration_num = item.split('_')[1]  # Get iteration number
                
                # Skip excluded iterations
                if exclude_iterations and int(iteration_num) in exclude_iterations:
                    logging.info(f"Skipping excluded iteration: {iteration_num}")
                    continue
                    
                logging.info(f"Processing iteration folder: {item}")
                
                if iteration_num not in iterations_data:
                    iterations_data[iteration_num] = {}
                
                iteration_folder = os.path.join(folder_path, item)
                
                # Process all annotation files in the iteration folder
                for file in os.listdir(iteration_folder):
                    if file.endswith('.xlsx'):
                        logging.info(f"Processing annotation file: {file}")
                        try:
                            # Extract annotator code from filename, joining all parts after iteration_X_ except the file extension
                            annotator = '_'.join(file.split('_')[2:]).split('.')[0]

                            if annotator != 'agreement' and annotator != 'discussion':
                                file_path = os.path.join(iteration_folder, file)
                                
                                # Use pandas to read Excel file instead of open()
                                df = pd.read_excel(file_path)
                                iterations_data[iteration_num][annotator] = parse_annotations(df)
                            
                        except Exception as e:
                            logging.error(f"Error reading file {file}: {str(e)}")
                
                if not iterations_data[iteration_num]:
                    logging.warning(f"No valid annotation files found in {iteration_folder}")
    
    except Exception as e:
        logging.error(f"Error in get_annotation_data: {str(e)}")
        raise
    
    logging.info(f"Completed processing. Found data for {len(iterations_data)} iterations")
    return iterations_data

def parse_annotations(df):
    """Parse annotation DataFrame into a matrix of 0s and 1s for each emotion."""
    logging.info("Starting to parse annotations")
    annotations = []
    emotions = ['Joy', 'Trust', 'Fear', 'Surprise', 'Sadness', 
                'Disgust', 'Anger', 'Anticipation', 'Neutral', 'Reject']
    
    try:
        for _, row in df.iterrows():
            # Convert values to 1 if present (1) and 0 if absent (0 or empty)
            emotion_vector = [1 if str(row.get(emotion, '')).strip() == '1' else 0 
                            for emotion in emotions]
            annotations.append(emotion_vector)
        
        logging.info(f"Successfully parsed {len(annotations)} annotations")
        return np.array(annotations)
    except Exception as e:
        logging.error(f"Error in parse_annotations: {str(e)}")
        raise

def calculate_pairwise_cohen_kappa(iteration, annotations, class_level=False):
    """Calculate Cohen's Kappa for each pair of annotators."""
    annotators = [ann for ann in all_annotators if ann in annotations]
    pair_kappas = {}
    class_kappas = {} if class_level else None

    logging.info(f"Iteration {iteration}. Valid annotators: {annotators}")
    
    if len(annotators) < 2:
        print(f"Warning: Not enough valid annotators to calculate pairwise kappa (found {len(annotators)})")
        print(annotators)
        return ({}, {}) if class_level else {}
    
    emotions = ['Joy', 'Trust', 'Fear', 'Surprise', 'Sadness', 
                'Disgust', 'Anger', 'Anticipation', 'Neutral', 'Reject']
    
    for i in range(len(annotators)):
        for j in range(i + 1, len(annotators)):
            ann1, ann2 = annotators[i], annotators[j]
            
            labels1 = annotations[ann1]
            labels2 = annotations[ann2]
            
            if len(labels1) == 0 or len(labels2) == 0:
                print(f"Warning: Empty labels found for {ann1}-{ann2}")
                continue
            
            # Calculate overall kappa
            labels1_flat = labels1.flatten()
            labels2_flat = labels2.flatten()
            kappa = cohen_kappa_score(labels1_flat, labels2_flat)
            pair_kappas[f"{ann1}-{ann2}"] = kappa
            
            # Calculate class-level kappa if requested
            if class_level:
                if f"{ann1}-{ann2}" not in class_kappas:
                    class_kappas[f"{ann1}-{ann2}"] = {}
                
                # Calculate kappa for each emotion separately
                for emotion_idx, emotion in enumerate(emotions):
                    try:
                        # Get labels for this emotion
                        emotion_labels1 = labels1[:, emotion_idx]
                        emotion_labels2 = labels2[:, emotion_idx]
                        
                        # Check if there's any variation in the labels
                        if len(np.unique(emotion_labels1)) > 1 or len(np.unique(emotion_labels2)) > 1:
                            emotion_kappa = cohen_kappa_score(emotion_labels1, emotion_labels2)
                            class_kappas[f"{ann1}-{ann2}"][emotion] = emotion_kappa
                        else:
                            # If all annotations are the same (all 0s or all 1s)
                            # Check if both annotators agree
                            if np.array_equal(emotion_labels1, emotion_labels2):
                                class_kappas[f"{ann1}-{ann2}"][emotion] = 1.0
                            else:
                                class_kappas[f"{ann1}-{ann2}"][emotion] = 0.0
                            
                    except Exception as e:
                        logging.warning(f"Could not calculate kappa for {emotion}: {str(e)}")
                        # Instead of None, use 0.0 for cases where kappa can't be calculated
                        class_kappas[f"{ann1}-{ann2}"][emotion] = 0.0
    
    return (pair_kappas, class_kappas) if class_level else pair_kappas

def create_agreement_heatmap(all_pair_kappas, all_annotators, output_path, show_full_matrix=True):
    """
    Create a heatmap of agreement scores.
    
    Args:
        all_pair_kappas: Dictionary of pairwise kappa scores
        all_annotators: List of annotator names
        output_path: Path to save the heatmap
        show_full_matrix: If False, only shows lower triangle of the matrix
    """
    # Create the matrix for the heatmap
    n = len(all_annotators)
    matrix = np.zeros((n, n))
    
    # Fill the matrix
    for i, ann1 in enumerate(all_annotators):
        for j, ann2 in enumerate(all_annotators):
            if ann1 == ann2:
                matrix[i, j] = 1.0
            else:
                # Try both orderings of the pair
                pair1 = f"{ann1}-{ann2}"
                pair2 = f"{ann2}-{ann1}"
                if pair1 in all_pair_kappas:
                    matrix[i, j] = np.mean(all_pair_kappas[pair1])
                elif pair2 in all_pair_kappas:
                    matrix[i, j] = np.mean(all_pair_kappas[pair2])

    # Create custom annotation function
    def format_value(val):
        if val in [0, 1]:
            return '-'
        return f'{val:.2f}'

    # Create mask for values that are 0 or 1
    mask = (matrix == 0) | (matrix == 1)
    
    # Add mask for upper triangle if not showing full matrix
    if not show_full_matrix:
        mask = mask | np.triu(np.ones_like(matrix), k=1).astype(bool)
    
    # Create the heatmap
    plt.figure(figsize=(12, 10))
    
    # Create base heatmap
    sns.heatmap(matrix, 
                annot=True,
                fmt='.2f',
                cmap='RdYlGn',
                xticklabels=heatmap_tags,
                yticklabels=heatmap_tags,
                vmin=0,
                vmax=1,
                square=True,
                mask=mask,
                cbar=True,
                annot_kws={'size': 16})

    # Overlay for special values
    sns.heatmap(matrix, 
                annot=[[format_value(val) for val in row] for row in matrix],
                fmt='',
                cmap=['#F5F5F5'],
                xticklabels=heatmap_tags,
                yticklabels=heatmap_tags,
                vmin=0,
                vmax=1,
                square=True,
                mask=~mask,
                cbar=False,
                annot_kws={'size': 16, 'color': '#A5A5A5'})

    plt.xticks(rotation=45, ha='right', fontsize=14)
    plt.yticks(rotation=0, fontsize=14)
    
    cbar = plt.gca().collections[0].colorbar
    cbar.ax.tick_labels = plt.setp(cbar.ax.get_yticklabels(), fontsize=12)

    plt.tight_layout()

    # Save the heatmap
    heatmap_path = output_path.replace('.xlsx', '_heatmap.png')
    plt.savefig(heatmap_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"Saved heatmap to: {heatmap_path}")

def create_excel_report(iterations_data, output_path, show_full_matrix=False, class_level=False):
    """Create Excel report with all statistics."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Agreement Statistics"
    
    # Add headers
    ws['A1'] = "Statistics Report"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    emotions = ['Joy', 'Trust', 'Fear', 'Surprise', 'Sadness', 
                'Disgust', 'Anger', 'Anticipation', 'Neutral', 'Reject']

    # Initialize dictionaries to store class-level statistics
    if class_level:
        all_class_kappas = {emotion: [] for emotion in emotions}

    # Process iterations
    for iteration, annotations in sorted(iterations_data.items(), key=lambda x: int(x[0])):
        ws[f'A{row}'] = f"Iteration {iteration}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Calculate kappas
        if class_level:
            pair_kappas, class_kappas = calculate_pairwise_cohen_kappa(iteration, annotations, class_level=True)
            # Just collect the class-level kappas for global statistics
            for pair_data in class_kappas.values():
                for emotion in emotions:
                    if pair_data[emotion] is not None:
                        all_class_kappas[emotion].append(pair_data[emotion])
        else:
            pair_kappas = calculate_pairwise_cohen_kappa(iteration, annotations)

        # Pairwise Cohen's Kappa
        ws[f'A{row}'] = "Pairwise Cohen's Kappa"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Create matrix headers using all possible annotators
        for i, ann in enumerate(all_annotators):
            ws.cell(row=row, column=i+2, value=ann)
            ws.cell(row=row+i+1, column=1, value=ann)
        
        # Create matrix of kappa values
        annotator_averages = {ann: [] for ann in all_annotators}
        for i, ann1 in enumerate(all_annotators):
            for j, ann2 in enumerate(all_annotators):
                if ann1 == ann2 and ann1 in annotations:
                    ws.cell(row=row+i+1, column=j+2, value=1.0)
                elif f"{ann1}-{ann2}" in pair_kappas:
                    value = pair_kappas[f"{ann1}-{ann2}"]
                    ws.cell(row=row+i+1, column=j+2, value=value)
                    annotator_averages[ann1].append(value)
                    annotator_averages[ann2].append(value)

        # Add "Average" row and column
        avg_row = row + len(all_annotators) + 1
        ws.cell(row=avg_row, column=1, value="Average")
        ws.cell(row=row, column=len(all_annotators)+2, value="Average")

        # Calculate and fill averages only for annotators with values
        all_values = []
        for i, ann in enumerate(all_annotators):
            if annotator_averages[ann]:  # Only calculate average if there are values
                ann_avg = np.mean(annotator_averages[ann])
                # Fill row average
                ws.cell(row=row+i+1, column=len(all_annotators)+2, value=ann_avg)
                # Fill column average
                ws.cell(row=avg_row, column=i+2, value=ann_avg)
                all_values.extend(annotator_averages[ann])

        # Calculate and fill total average
        if all_values:  # Only calculate if there are values
            total_avg = np.mean(all_values)
            ws.cell(row=avg_row, column=len(all_annotators)+2, value=total_avg)

        row = avg_row + 2

    # Add new section for overall pairwise agreements
    ws[f'A{row}'] = "Overall Average Pairwise Agreements"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    # Headers for the simple table
    ws[f'A{row}'] = "Annotator Pair"
    ws[f'B{row}'] = "Average Agreement"
    ws[f'C{row}'] = "Number of Iterations"
    row += 1

    # Initialize dictionary to store all pairwise kappas
    all_pair_kappas = {}

    # Collect all pairwise kappas across iterations
    for iteration, annotations in iterations_data.items():
        pair_kappas = calculate_pairwise_cohen_kappa(iteration, annotations)
        for pair, kappa in pair_kappas.items():
            if pair not in all_pair_kappas:
                all_pair_kappas[pair] = []
            all_pair_kappas[pair].append(kappa)

    # Sort pairs by average agreement
    pair_averages = {
        pair: (np.mean(kappas), len(kappas))
        for pair, kappas in all_pair_kappas.items()
    }
    sorted_pairs = sorted(pair_averages.items(), key=lambda x: x[1][0], reverse=True)

    # Write the pairs and their averages
    for pair, (avg_kappa, num_iterations) in sorted_pairs:
        ws[f'A{row}'] = pair
        ws[f'B{row}'] = f"{avg_kappa:.3f}"
        ws[f'C{row}'] = num_iterations
        row += 1

    # After writing the flat list of pairs, add a matrix view
    row += 2
    ws[f'A{row}'] = "Overall Average Agreements Matrix"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    matrix_start_row = row  # Save the starting row of the matrix

    # Create matrix headers
    for i, ann in enumerate(all_annotators):
        ws.cell(row=row, column=i+2, value=ann)
        ws.cell(row=row+i+1, column=1, value=ann)

    # Fill the matrix with average kappa values
    for i, ann1 in enumerate(all_annotators):
        for j, ann2 in enumerate(all_annotators):
            if ann1 == ann2:
                ws.cell(row=row+i+1, column=j+2, value=1.0)
            else:
                # Try both orderings of the pair
                pair1 = f"{ann1}-{ann2}"
                pair2 = f"{ann2}-{ann1}"
                if pair1 in all_pair_kappas:
                    value = np.mean(all_pair_kappas[pair1])
                    ws.cell(row=row+i+1, column=j+2, value=value)
                elif pair2 in all_pair_kappas:
                    value = np.mean(all_pair_kappas[pair2])
                    ws.cell(row=row+i+1, column=j+2, value=value)

    # Update row to be after the matrix
    row = matrix_start_row + len(all_annotators) + 2

    # Add global class-level statistics if requested (only at the end)
    if class_level:
        row += 2
        ws[f'A{row}'] = "Global Class-level Statistics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Emotion"
        ws[f'B{row}'] = "Average Kappa"
        ws[f'C{row}'] = "Number of Valid Pairs"
        row += 1
        
        # Sort emotions by average kappa for better readability
        emotion_stats = []
        for emotion in emotions:
            values = [v for v in all_class_kappas[emotion] if v is not None]
            if values:
                avg_kappa = np.mean(values)
                emotion_stats.append((emotion, avg_kappa, len(values)))
        
        # Sort by average kappa in descending order
        emotion_stats.sort(key=lambda x: x[1], reverse=True)
        
        # Write sorted statistics
        for emotion, avg_kappa, valid_pairs in emotion_stats:
            ws[f'A{row}'] = emotion
            ws[f'B{row}'] = f"{avg_kappa:.3f}"
            ws[f'C{row}'] = valid_pairs
            row += 1

    wb.save(output_path)
    
    # Create and save the heatmap
    create_agreement_heatmap(all_pair_kappas, all_annotators, output_path, show_full_matrix=show_full_matrix)

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Calculate agreement statistics for annotations.')
    parser.add_argument('--input_folder', 
                       help='Path to the folder containing annotation data')
    parser.add_argument('--output_folder',
                       help='Path to the folder where output files will be saved',
                       default=None)
    parser.add_argument('--exclude', 
                       help='Comma-separated list of iterations to exclude (e.g., "0,1,2")',
                       default=None)
    parser.add_argument('--full_matrix', 
                       action='store_true',
                       help='Show full agreement matrix instead of just lower triangle')
    parser.add_argument('--annotators',
                       help='Comma-separated list of annotators to analyze',
                       default=None)
    parser.add_argument('--class-level', 
                       action='store_true',
                       help='Calculate class-level (per-label) agreement statistics')

    args = parser.parse_args()
    
    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('annotation_processing.log'),
            logging.StreamHandler()
        ]
    )
    
    # Parse excluded iterations if provided
    exclude_iterations = None
    if args.exclude:
        try:
            exclude_iterations = [int(x) for x in args.exclude.split(',')]
            logging.info(f"Excluding iterations: {exclude_iterations}")
        except ValueError:
            logging.error("Invalid format for excluded iterations. Use comma-separated integers.")
            print("Error: Invalid format for excluded iterations. Use comma-separated integers (e.g., 0,1,2)")
            sys.exit(1)
    
    # Parse annotators if provided
    global all_annotators, heatmap_tags
    if args.annotators:
        all_annotators = args.annotators.split(',')
        # Update heatmap_tags based on number of annotators provided
        heatmap_tags = [f'Annotator_{i+1}' for i in range(len(all_annotators))]
        logging.info(f"Using provided annotators: {all_annotators}")
    
    if not os.path.exists(args.input_folder):
        logging.error(f"Folder does not exist: {args.input_folder}")
        print(f"Error: Folder '{args.input_folder}' does not exist.")
        sys.exit(1)
    
    # Handle output folder
    output_folder = args.output_folder if args.output_folder else args.input_folder
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        logging.info(f"Created output folder: {output_folder}")
    
    try:
        # Process all iterations
        logging.info("Starting to process iterations")
        iterations_data = get_annotation_data(args.input_folder, exclude_iterations)
        
        # Create and save report
        annotators_str = '-'.join(all_annotators)  # Join annotator names with hyphens
        proposed_filename = f'agreement-statistics-{annotators_str}.xlsx'
        
        # Windows max path length is 260, and max filename is 255
        if len(proposed_filename) > 255:
            output_filename = 'agreement-statistics-ALL.xlsx'
        else:
            output_filename = proposed_filename
            
        output_path = os.path.join(output_folder, output_filename)
        logging.info(f"Creating Excel report at: {output_path}")
        
        create_excel_report(iterations_data, output_path, args.full_matrix, args.class_level)
        logging.info("Processing completed successfully")
        print(f"Agreement statistics have been saved to: {output_path}")
    
    except Exception as e:
        logging.error(f"Fatal error in main execution: {str(e)}")
        raise

if __name__ == "__main__":
    main()
