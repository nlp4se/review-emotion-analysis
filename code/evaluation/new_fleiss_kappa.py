import os
import sys
import pandas as pd
import numpy as np
from sklearn.metrics import cohen_kappa_score
from openpyxl import Workbook
from openpyxl.styles import Font
import logging

def get_annotation_data(folder_path):
    """Extract annotation data from all iteration folders."""
    logging.info(f"Starting to process folder: {folder_path}")
    iterations_data = {}
    
    try:
        # Find all iteration folders
        for item in os.listdir(folder_path):
            if item.startswith('iteration_'):
                iteration_num = item.split('_')[1]  # Get iteration number
                logging.info(f"Processing iteration folder: {item}")
                
                if iteration_num not in iterations_data:
                    iterations_data[iteration_num] = {}
                
                iteration_folder = os.path.join(folder_path, item)
                
                # Process all annotation files in the iteration folder
                for file in os.listdir(iteration_folder):
                    if file.endswith('.xlsx'):
                        logging.info(f"Processing annotation file: {file}")
                        try:
                            # Extract annotator code from filename (e.g., 'QM' from 'iteration_0_QM.xlsx')
                            annotator = file.split('_')[2].split('.')[0]

                            if len(annotator) == 2:
                                file_path = os.path.join(iteration_folder, file)
                                
                                # Use pandas to read Excel file instead of open()
                                df = pd.read_excel(file_path)
                                iterations_data[iteration_num][annotator] = parse_annotations(df)
                            
                        except Exception as e:
                            logging.error(f"Error reading file {file}: {str(e)}")
                
                if not iterations_data[iteration_num]:
                    logging.warning(f"No valid annotation files found in {iteration_folder}")
    
    except Exception as e:
        logging.error(f"Error in get_annotation_data: {str(e)}")
        raise
    
    logging.info(f"Completed processing. Found data for {len(iterations_data)} iterations")
    return iterations_data

def parse_annotations(df):
    """Parse annotation DataFrame into a matrix of 0s and 1s for each emotion."""
    logging.info("Starting to parse annotations")
    annotations = []
    emotions = ['Joy', 'Trust', 'Fear', 'Surprise', 'Sadness', 
                'Disgust', 'Anger', 'Anticipation', 'Neutral', 'Reject']
    
    try:
        for _, row in df.iterrows():
            # Convert values to 1 if present (1) and 0 if absent (0 or empty)
            emotion_vector = [1 if str(row.get(emotion, '')).strip() == '1' else 0 
                            for emotion in emotions]
            annotations.append(emotion_vector)
        
        logging.info(f"Successfully parsed {len(annotations)} annotations")
        return np.array(annotations)
    except Exception as e:
        logging.error(f"Error in parse_annotations: {str(e)}")
        raise

def calculate_pairwise_cohen_kappa(annotations):
    """Calculate Cohen's Kappa for each pair of annotators."""
    annotators = list(annotations.keys())
    pair_kappas = {}
    
    if len(annotators) < 2:
        print(f"Warning: Not enough annotators to calculate pairwise kappa (found {len(annotators)})")
        return {}
    
    logging.info(f"Annotators: {annotators}")
    for i in range(len(annotators)):
        for j in range(i + 1, len(annotators)):
            ann1, ann2 = annotators[i], annotators[j]
            
            # Get full annotation matrices for both annotators
            labels1 = annotations[ann1]
            labels2 = annotations[ann2]
            
            # Add defensive check
            if len(labels1) == 0 or len(labels2) == 0:
                print(f"Warning: Empty labels found for {ann1}-{ann2}")
                continue
            
            # Flatten the matrices into 1D arrays
            labels1_flat = labels1.flatten()
            labels2_flat = labels2.flatten()
            
            # Calculate kappa on all binary decisions
            kappa = cohen_kappa_score(labels1_flat, labels2_flat)
            pair_kappas[f"{ann1}-{ann2}"] = kappa
    
    return pair_kappas

def calculate_fleiss_kappa_for_iteration(annotations):
    """Calculate Fleiss' Kappa for all annotators in an iteration."""
    # Add defensive check
    if not annotations:
        print("Warning: No annotations found for this iteration")
        return float('nan')
        
    num_reviews = len(next(iter(annotations.values())))
    num_emotions = 10
    num_annotators = len(annotations)
    
    # Add defensive check
    if num_reviews == 0:
        print("Warning: No reviews found in annotations")
        return float('nan')
    
    # Create a matrix where each row represents one item (review-emotion pair)
    # and contains the number of raters who assigned each category (0 or 1)
    n = num_annotators  # number of raters
    N = num_reviews * num_emotions  # number of subjects (review-emotion pairs)
    
    # Initialize ratings matrix for each subject
    ratings = np.zeros((N, 2))  # 2 categories: 0 and 1
    
    # Fill the ratings matrix
    for i in range(num_reviews):
        for j in range(num_emotions):
            idx = i * num_emotions + j
            # Count ratings for this review-emotion pair
            count_ones = sum(1 for ann in annotations.values() 
                           if ann[i][j] == 1)
            ratings[idx, 1] = count_ones  # number of 1s
            ratings[idx, 0] = n - count_ones  # number of 0s
    
    # Calculate P_i (proportion of agreement for each subject)
    P_i = np.sum(ratings * (ratings - 1), axis=1) / (n * (n - 1))
    P_bar = np.mean(P_i)  # mean agreement across subjects
    
    # Calculate P_e (expected agreement by chance)
    p_j = np.sum(ratings, axis=0) / (N * n)  # proportion of assignments in each category
    P_e = np.sum(p_j ** 2)
    
    # Calculate Fleiss' Kappa
    kappa = (P_bar - P_e) / (1 - P_e)
    
    return kappa

def create_excel_report(iterations_data, output_path):
    """Create Excel report with all statistics."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Agreement Statistics"
    
    # Add headers
    ws['A1'] = "Statistics Report"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    # Fixed list of all possible annotators in specified order
    all_annotators = ['QM', 'MT', 'MO', 'JM', 'XF', 'GPT-4o', 'GPT-4o-mini']

    # Convert iteration keys to integers and sort numerically
    for iteration, annotations in sorted(iterations_data.items(), key=lambda x: int(x[0])):
        ws[f'A{row}'] = f"Iteration {iteration}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Pairwise Cohen's Kappa
        pair_kappas = calculate_pairwise_cohen_kappa(annotations)
        ws[f'A{row}'] = "Pairwise Cohen's Kappa"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Create matrix headers using all possible annotators
        for i, ann in enumerate(all_annotators):
            ws.cell(row=row, column=i+2, value=ann)
            ws.cell(row=row+i+1, column=1, value=ann)
        
        # Create matrix of kappa values
        annotator_averages = {ann: [] for ann in all_annotators}
        for i, ann1 in enumerate(all_annotators):
            for j, ann2 in enumerate(all_annotators):
                if ann1 == ann2 and ann1 in annotations:
                    ws.cell(row=row+i+1, column=j+2, value=1.0)
                elif f"{ann1}-{ann2}" in pair_kappas:
                    value = pair_kappas[f"{ann1}-{ann2}"]
                    ws.cell(row=row+i+1, column=j+2, value=value)
                    annotator_averages[ann1].append(value)
                    annotator_averages[ann2].append(value)
                # If no value exists, cell remains empty

        # Add "Average" row and column
        avg_row = row + len(all_annotators) + 1
        ws.cell(row=avg_row, column=1, value="Average")
        ws.cell(row=row, column=len(all_annotators)+2, value="Average")

        # Calculate and fill averages only for annotators with values
        all_values = []
        for i, ann in enumerate(all_annotators):
            if annotator_averages[ann]:  # Only calculate average if there are values
                ann_avg = np.mean(annotator_averages[ann])
                # Fill row average
                ws.cell(row=row+i+1, column=len(all_annotators)+2, value=ann_avg)
                # Fill column average
                ws.cell(row=avg_row, column=i+2, value=ann_avg)
                all_values.extend(annotator_averages[ann])

        # Calculate and fill total average
        if all_values:  # Only calculate if there are values
            total_avg = np.mean(all_values)
            ws.cell(row=avg_row, column=len(all_annotators)+2, value=total_avg)

        row = avg_row + 2

        # Fleiss' Kappa for iteration
        fleiss = calculate_fleiss_kappa_for_iteration(annotations)
        ws[f'A{row}'] = f"Fleiss' Kappa for iteration {iteration}: {fleiss:.3f}"
        row += 2

    # Overall statistics
    ws[f'A{row}'] = "Overall Statistics"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    overall_fleiss_kappas = [calculate_fleiss_kappa_for_iteration(annotations) for annotations in iterations_data.values()]
    overall_fleiss = np.mean(overall_fleiss_kappas)
    ws[f'A{row}'] = f"Average Fleiss' Kappa across all iterations: {overall_fleiss:.3f}"
    
    wb.save(output_path)

def main():
    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('annotation_processing.log'),
            logging.StreamHandler()
        ]
    )
    
    if len(sys.argv) != 2:
        logging.error("Invalid number of arguments")
        print("Usage: python new_fleiss_kappa.py <folder_path>")
        sys.exit(1)
    
    folder_path = sys.argv[1]
    if not os.path.exists(folder_path):
        logging.error(f"Folder does not exist: {folder_path}")
        print(f"Error: Folder '{folder_path}' does not exist.")
        sys.exit(1)
    
    try:
        # Process all iterations
        logging.info("Starting to process iterations")
        iterations_data = get_annotation_data(folder_path)
        
        # Create and save report
        output_path = os.path.join(folder_path, 'agreement-statistics.xlsx')
        logging.info(f"Creating Excel report at: {output_path}")
        create_excel_report(iterations_data, output_path)
        logging.info("Processing completed successfully")
        print(f"Agreement statistics have been saved to: {output_path}")
    
    except Exception as e:
        logging.error(f"Fatal error in main execution: {str(e)}")
        raise

if __name__ == "__main__":
    main()
