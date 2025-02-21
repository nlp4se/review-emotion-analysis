import os
import sys
import pandas as pd
import numpy as np
from sklearn.metrics import cohen_kappa_score
from openpyxl import Workbook
from openpyxl.styles import Font
import logging
import seaborn as sns
import matplotlib.pyplot as plt

# Fixed list of all possible annotators in specified order
#all_annotators = ['QM', 'MT', 'MO', 'JM', 'XF', 'gpt-4o', 'gpt-4o-mini']
#all_annotators = ['QM', 'MT', 'MO', 'JM', 'XF', 'gpt-4o-mini']
#all_annotators = ['QM', 'MT', 'MO', 'JM', 'XF', 'gpt-4o']
#all_annotators = ['QM', 'MT', 'MO', 'JM', 'XF', 'mistral-large-2411']
#all_annotators = ['QM', 'MT', 'MO', 'JM', 'XF', 'gemini-2-0-flash']
#all_annotators = ['QM', 'MT', 'MO', 'JM', 'XF']
#all_annotators = ['gpt-4o', 'mistral-large-2411', 'gemini-2-0-flash']
#all_annotators = ['gpt-4o-1','gpt-4o-2','gpt-4o-3']
#all_annotators = ['mistral-large-2411-1','mistral-large-2411-2','mistral-large-2411-3']
#all_annotators = ['gemini-2-0-flash-1','gemini-2-0-flash-2','gemini-2-0-flash-3']
all_annotators = ['QM', 'MT', 'MO', 'JM', 'XF', 
                  'agreement_JM-MO-MT-QM-XF', 
                  'agreement_gpt-4o-1-gpt-4o-2-gpt-4o-3',
                  'agreement_mistral-large-2411-1-mistral-large-2411-2-mistral-large-2411-3',
                  'agreement_gemini-2-0-flash-1-gemini-2-0-flash-2-gemini-2-0-flash-3', 
                  'agreement_gemini-2-0-flash-1-gemini-2-0-flash-2-gemini-2-0-flash-3-gpt-4o-1-gpt-4o-2-gpt-4o-3-mistral-large-2411-1-mistral-large-2411-2-mistral-large-2411-3']
heatmap_tags = ['Human_1', 'Human_2', 'Human_3', 'Human_4', 'Human_5', 'Human_A', 'GPT_A', 'Mistral_A', 'Gemini_A', 'LLM_A']

def get_annotation_data(folder_path, exclude_iterations=None):
    """Extract annotation data from all iteration folders."""
    logging.info(f"Starting to process folder: {folder_path}")
    iterations_data = {}
    
    try:
        # Find all iteration folders
        for item in os.listdir(folder_path):
            if item.startswith('iteration_'):
                iteration_num = item.split('_')[1]  # Get iteration number
                
                # Skip excluded iterations
                if exclude_iterations and int(iteration_num) in exclude_iterations:
                    logging.info(f"Skipping excluded iteration: {iteration_num}")
                    continue
                    
                logging.info(f"Processing iteration folder: {item}")
                
                if iteration_num not in iterations_data:
                    iterations_data[iteration_num] = {}
                
                iteration_folder = os.path.join(folder_path, item)
                
                # Process all annotation files in the iteration folder
                for file in os.listdir(iteration_folder):
                    if file.endswith('.xlsx'):
                        logging.info(f"Processing annotation file: {file}")
                        try:
                            # Extract annotator code from filename, joining all parts after iteration_X_ except the file extension
                            annotator = '_'.join(file.split('_')[2:]).split('.')[0]

                            if annotator != 'agreement' and annotator != 'discussion':
                                file_path = os.path.join(iteration_folder, file)
                                
                                # Use pandas to read Excel file instead of open()
                                df = pd.read_excel(file_path)
                                iterations_data[iteration_num][annotator] = parse_annotations(df)
                            
                        except Exception as e:
                            logging.error(f"Error reading file {file}: {str(e)}")
                
                if not iterations_data[iteration_num]:
                    logging.warning(f"No valid annotation files found in {iteration_folder}")
    
    except Exception as e:
        logging.error(f"Error in get_annotation_data: {str(e)}")
        raise
    
    logging.info(f"Completed processing. Found data for {len(iterations_data)} iterations")
    return iterations_data

def parse_annotations(df):
    """Parse annotation DataFrame into a matrix of 0s and 1s for each emotion."""
    logging.info("Starting to parse annotations")
    annotations = []
    emotions = ['Joy', 'Trust', 'Fear', 'Surprise', 'Sadness', 
                'Disgust', 'Anger', 'Anticipation', 'Neutral', 'Reject']
    
    try:
        for _, row in df.iterrows():
            # Convert values to 1 if present (1) and 0 if absent (0 or empty)
            emotion_vector = [1 if str(row.get(emotion, '')).strip() == '1' else 0 
                            for emotion in emotions]
            annotations.append(emotion_vector)
        
        logging.info(f"Successfully parsed {len(annotations)} annotations")
        return np.array(annotations)
    except Exception as e:
        logging.error(f"Error in parse_annotations: {str(e)}")
        raise

def calculate_pairwise_cohen_kappa(iteration, annotations):
    """Calculate Cohen's Kappa for each pair of annotators."""
    # Only consider annotators that are in our predefined list
    annotators = [ann for ann in all_annotators if ann in annotations]
    pair_kappas = {}

    logging.info(f"Iteration {iteration}. Valid annotators: {annotators}")
    
    if len(annotators) < 2:
        print(f"Warning: Not enough valid annotators to calculate pairwise kappa (found {len(annotators)})")
        print(annotators)
        return {}
    
    for i in range(len(annotators)):
        for j in range(i + 1, len(annotators)):
            ann1, ann2 = annotators[i], annotators[j]
            
            # Get full annotation matrices for both annotators
            labels1 = annotations[ann1]
            labels2 = annotations[ann2]
            
            # Add defensive check
            if len(labels1) == 0 or len(labels2) == 0:
                print(f"Warning: Empty labels found for {ann1}-{ann2}")
                continue
            
            # Flatten the matrices into 1D arrays
            labels1_flat = labels1.flatten()
            labels2_flat = labels2.flatten()
            
            # Calculate kappa on all binary decisions
            kappa = cohen_kappa_score(labels1_flat, labels2_flat)
            pair_kappas[f"{ann1}-{ann2}"] = kappa
    
    return pair_kappas

def calculate_fleiss_kappa_for_iteration(annotations):
    """Calculate Fleiss' Kappa for all annotators in an iteration."""
    # Filter annotations to only include predefined annotators
    filtered_annotations = {ann: data for ann, data in annotations.items() 
                          if ann in all_annotators}
    
    if not filtered_annotations:
        print("Warning: No valid annotations found for this iteration")
        return float('nan')
        
    num_reviews = len(next(iter(filtered_annotations.values())))
    num_emotions = 10  # Number of emotions being rated
    num_annotators = len(filtered_annotations)
    
    if num_reviews == 0:
        print("Warning: No reviews found in annotations")
        return float('nan')

    # For each review-emotion pair, we need to count how many annotators marked 0 and 1
    # Create a matrix of shape (num_reviews * num_emotions, 2) where
    # each row represents a review-emotion pair and columns represent counts of 0s and 1s
    ratings_matrix = np.zeros((num_reviews * num_emotions, 2))
    
    # Fill the ratings matrix
    for i in range(num_reviews):
        for j in range(num_emotions):
            idx = i * num_emotions + j
            # Count how many annotators marked this emotion as present (1)
            count_ones = sum(1 for ann in filtered_annotations.values() 
                           if ann[i][j] == 1)
            ratings_matrix[idx, 1] = count_ones
            ratings_matrix[idx, 0] = num_annotators - count_ones
    
    # Calculate P_i (proportion of agreeing pairs out of all possible pairs)
    # For each item (review-emotion pair)
    P_i = np.sum(ratings_matrix * (ratings_matrix - 1), axis=1) / (num_annotators * (num_annotators - 1))
    
    # Calculate P̄ (mean agreement across all items)
    P_bar = np.mean(P_i)
    
    # Calculate P_e (expected agreement by chance)
    # First get the proportion of all ratings in each category (0 and 1)
    p_j = np.sum(ratings_matrix, axis=0) / (ratings_matrix.shape[0] * num_annotators)
    P_e = np.sum(p_j ** 2)
    
    # Calculate Fleiss' Kappa
    kappa = (P_bar - P_e) / (1 - P_e)
    
    return kappa

def create_agreement_heatmap(all_pair_kappas, all_annotators, output_path, show_full_matrix=True):
    """
    Create a heatmap of agreement scores.
    
    Args:
        all_pair_kappas: Dictionary of pairwise kappa scores
        all_annotators: List of annotator names
        output_path: Path to save the heatmap
        show_full_matrix: If False, only shows lower triangle of the matrix
    """
    # Create the matrix for the heatmap
    n = len(all_annotators)
    matrix = np.zeros((n, n))
    
    # Fill the matrix
    for i, ann1 in enumerate(all_annotators):
        for j, ann2 in enumerate(all_annotators):
            if ann1 == ann2:
                matrix[i, j] = 1.0
            else:
                # Try both orderings of the pair
                pair1 = f"{ann1}-{ann2}"
                pair2 = f"{ann2}-{ann1}"
                if pair1 in all_pair_kappas:
                    matrix[i, j] = np.mean(all_pair_kappas[pair1])
                elif pair2 in all_pair_kappas:
                    matrix[i, j] = np.mean(all_pair_kappas[pair2])

    # Create custom annotation function
    def format_value(val):
        if val in [0, 1]:
            return '-'
        return f'{val:.2f}'

    # Create mask for values that are 0 or 1
    mask = (matrix == 0) | (matrix == 1)
    
    # Add mask for upper triangle if not showing full matrix
    if not show_full_matrix:
        mask = mask | np.triu(np.ones_like(matrix), k=1).astype(bool)
    
    # Create the heatmap
    plt.figure(figsize=(12, 10))
    
    # Create base heatmap
    sns.heatmap(matrix, 
                annot=True,
                fmt='.2f',
                cmap='RdYlGn',
                xticklabels=heatmap_tags,
                yticklabels=heatmap_tags,
                vmin=0,
                vmax=1,
                square=True,
                mask=mask,
                cbar=True,
                annot_kws={'size': 16})

    # Overlay for special values
    sns.heatmap(matrix, 
                annot=[[format_value(val) for val in row] for row in matrix],
                fmt='',
                cmap=['#F5F5F5'],
                xticklabels=heatmap_tags,
                yticklabels=heatmap_tags,
                vmin=0,
                vmax=1,
                square=True,
                mask=~mask,
                cbar=False,
                annot_kws={'size': 16, 'color': '#A5A5A5'})

    plt.xticks(rotation=45, ha='right', fontsize=14)
    plt.yticks(rotation=0, fontsize=14)
    
    cbar = plt.gca().collections[0].colorbar
    cbar.ax.tick_labels = plt.setp(cbar.ax.get_yticklabels(), fontsize=12)

    plt.tight_layout()
    
    # Save the heatmap
    heatmap_path = output_path.replace('.xlsx', '_heatmap.png')
    plt.savefig(heatmap_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"Saved heatmap to: {heatmap_path}")

def create_excel_report(iterations_data, output_path, show_full_matrix=False):
    """Create Excel report with all statistics."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Agreement Statistics"
    
    # Add headers
    ws['A1'] = "Statistics Report"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3

    # Convert iteration keys to integers and sort numerically
    for iteration, annotations in sorted(iterations_data.items(), key=lambda x: int(x[0])):
        ws[f'A{row}'] = f"Iteration {iteration}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Pairwise Cohen's Kappa
        pair_kappas = calculate_pairwise_cohen_kappa(iteration, annotations)
        ws[f'A{row}'] = "Pairwise Cohen's Kappa"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Create matrix headers using all possible annotators
        for i, ann in enumerate(all_annotators):
            ws.cell(row=row, column=i+2, value=ann)
            ws.cell(row=row+i+1, column=1, value=ann)
        
        # Create matrix of kappa values
        annotator_averages = {ann: [] for ann in all_annotators}
        for i, ann1 in enumerate(all_annotators):
            for j, ann2 in enumerate(all_annotators):
                if ann1 == ann2 and ann1 in annotations:
                    ws.cell(row=row+i+1, column=j+2, value=1.0)
                elif f"{ann1}-{ann2}" in pair_kappas:
                    value = pair_kappas[f"{ann1}-{ann2}"]
                    ws.cell(row=row+i+1, column=j+2, value=value)
                    annotator_averages[ann1].append(value)
                    annotator_averages[ann2].append(value)
                # If no value exists, cell remains empty

        # Add "Average" row and column
        avg_row = row + len(all_annotators) + 1
        ws.cell(row=avg_row, column=1, value="Average")
        ws.cell(row=row, column=len(all_annotators)+2, value="Average")

        # Calculate and fill averages only for annotators with values
        all_values = []
        for i, ann in enumerate(all_annotators):
            if annotator_averages[ann]:  # Only calculate average if there are values
                ann_avg = np.mean(annotator_averages[ann])
                # Fill row average
                ws.cell(row=row+i+1, column=len(all_annotators)+2, value=ann_avg)
                # Fill column average
                ws.cell(row=avg_row, column=i+2, value=ann_avg)
                all_values.extend(annotator_averages[ann])

        # Calculate and fill total average
        if all_values:  # Only calculate if there are values
            total_avg = np.mean(all_values)
            ws.cell(row=avg_row, column=len(all_annotators)+2, value=total_avg)

        row = avg_row + 2

        # Fleiss' Kappa for iteration
        fleiss = calculate_fleiss_kappa_for_iteration(annotations)
        ws[f'A{row}'] = f"Fleiss' Kappa for iteration {iteration}: {fleiss:.3f}"
        row += 2

    # Overall statistics
    ws[f'A{row}'] = "Overall Statistics"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    overall_fleiss_kappas = [calculate_fleiss_kappa_for_iteration(annotations) for annotations in iterations_data.values()]
    overall_fleiss = np.mean(overall_fleiss_kappas)
    ws[f'A{row}'] = f"Average Fleiss' Kappa across all iterations: {overall_fleiss:.3f}"
    row += 2

    # Add new section for overall pairwise agreements
    ws[f'A{row}'] = "Overall Average Pairwise Agreements"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    # Headers for the simple table
    ws[f'A{row}'] = "Annotator Pair"
    ws[f'B{row}'] = "Average Agreement"
    ws[f'C{row}'] = "Number of Iterations"
    row += 1

    # Initialize dictionary to store all pairwise kappas
    all_pair_kappas = {}

    # Collect all pairwise kappas across iterations
    for iteration, annotations in iterations_data.items():
        pair_kappas = calculate_pairwise_cohen_kappa(iteration, annotations)
        for pair, kappa in pair_kappas.items():
            if pair not in all_pair_kappas:
                all_pair_kappas[pair] = []
            all_pair_kappas[pair].append(kappa)

    # Sort pairs by average agreement
    pair_averages = {
        pair: (np.mean(kappas), len(kappas))
        for pair, kappas in all_pair_kappas.items()
    }
    sorted_pairs = sorted(pair_averages.items(), key=lambda x: x[1][0], reverse=True)

    # Write the pairs and their averages
    for pair, (avg_kappa, num_iterations) in sorted_pairs:
        ws[f'A{row}'] = pair
        ws[f'B{row}'] = f"{avg_kappa:.3f}"
        ws[f'C{row}'] = num_iterations
        row += 1

    # After writing the flat list of pairs, add a matrix view
    row += 2
    ws[f'A{row}'] = "Overall Average Agreements Matrix"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    # Create matrix headers
    for i, ann in enumerate(all_annotators):
        ws.cell(row=row, column=i+2, value=ann)
        ws.cell(row=row+i+1, column=1, value=ann)

    # Fill the matrix with average kappa values
    for i, ann1 in enumerate(all_annotators):
        for j, ann2 in enumerate(all_annotators):
            if ann1 == ann2:
                ws.cell(row=row+i+1, column=j+2, value=1.0)
            else:
                # Try both orderings of the pair
                pair1 = f"{ann1}-{ann2}"
                pair2 = f"{ann2}-{ann1}"
                if pair1 in all_pair_kappas:
                    value = np.mean(all_pair_kappas[pair1])
                    ws.cell(row=row+i+1, column=j+2, value=value)
                elif pair2 in all_pair_kappas:
                    value = np.mean(all_pair_kappas[pair2])
                    ws.cell(row=row+i+1, column=j+2, value=value)

    wb.save(output_path)
    
    # Create and save the heatmap
    create_agreement_heatmap(all_pair_kappas, all_annotators, output_path, show_full_matrix=show_full_matrix)

def main():
    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('annotation_processing.log'),
            logging.StreamHandler()
        ]
    )
    
    if len(sys.argv) < 2:
        logging.error("Invalid number of arguments")
        print("Usage: python kappa_analysis.py <folder_path> [excluded_iterations] [show_full_matrix]")
        print("Example: python kappa_analysis.py ./data 0,1,2 true")
        sys.exit(1)
    
    folder_path = sys.argv[1]
    
    # Parse excluded iterations if provided
    exclude_iterations = None
    if len(sys.argv) > 2 and sys.argv[2] != 'true' and sys.argv[2] != 'false':
        try:
            exclude_iterations = [int(x) for x in sys.argv[2].split(',')]
            logging.info(f"Excluding iterations: {exclude_iterations}")
        except ValueError:
            logging.error("Invalid format for excluded iterations. Use comma-separated integers.")
            print("Error: Invalid format for excluded iterations. Use comma-separated integers (e.g., 0,1,2)")
            sys.exit(1)
    
    # Parse show_full_matrix parameter if provided
    show_full_matrix = False  # default value
    if len(sys.argv) > 3:
        show_full_matrix = sys.argv[3].lower() == 'true'
    elif len(sys.argv) == 3 and (sys.argv[2] == 'true' or sys.argv[2] == 'false'):
        show_full_matrix = sys.argv[2].lower() == 'true'
    
    if not os.path.exists(folder_path):
        logging.error(f"Folder does not exist: {folder_path}")
        print(f"Error: Folder '{folder_path}' does not exist.")
        sys.exit(1)
    
    try:
        # Process all iterations
        logging.info("Starting to process iterations")
        iterations_data = get_annotation_data(folder_path, exclude_iterations)
        
        # Create and save report
        annotators_str = '-'.join(all_annotators)  # Join annotator names with hyphens
        proposed_filename = f'agreement-statistics-{annotators_str}.xlsx'
        
        # Windows max path length is 260, and max filename is 255
        if len(proposed_filename) > 255:
            output_filename = 'agreement-statistics-ALL.xlsx'
        else:
            output_filename = proposed_filename
            
        output_path = os.path.join(folder_path, output_filename)
        logging.info(f"Creating Excel report at: {output_path}")
        
        # Pass show_full_matrix parameter to create_excel_report
        create_excel_report(iterations_data, output_path, show_full_matrix)
        logging.info("Processing completed successfully")
        print(f"Agreement statistics have been saved to: {output_path}")
    
    except Exception as e:
        logging.error(f"Fatal error in main execution: {str(e)}")
        raise

if __name__ == "__main__":
    main()
